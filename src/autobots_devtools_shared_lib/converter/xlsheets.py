"""
ExcelSheetsManager - A comprehensive wrapper for Excel file operations
with pandas DataFrame integration, using a file server backend.
"""

import io
import os
from pathlib import Path
from typing import List, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
from autobots_devtools_shared_lib.common.observability.logging_utils import (
    get_logger,
    set_conversation_id,
    setup_logging,
)

logger = get_logger(__name__)


class LocalFileClient:
    """
    Client for interacting with the local file system.

    Provides methods to read, write, and list files on the local filesystem.
    """

    def __init__(self, base_dir: str = "."):
        """
        Initialize the local file client.

        Args:
            base_dir: Base directory for file operations (default: current directory)
        """
        self.base_dir = base_dir

    def _get_base_path(
        self, user_name: str | None = None, repo_name: str | None = None, jira_number: str | None = None
    ) -> Path:
        """
        Get the base path for file operations.

        Args:
            user_name: User name (not used in local filesystem)
            repo_name: Repository name (not used in local filesystem)
            jira_number: JIRA number (not used in local filesystem)
        """
        # Optional workspace components: either all three or none
        components = [user_name, repo_name, jira_number]
        provided = [c for c in components if c]
        if provided and len(provided) != 3:
            logger.error(
                "Partial workspace components provided",
            )
            raise ValueError(
                "Either provide all of user_name, repo_name, jira_number or none of them"
            )

        base_path = Path(self.base_dir)
        logger.info(f"Base path resolved to {str(base_path)}")

        if all(components):
            # mypy/pyright: components are all non-None here
            assert user_name is not None and repo_name is not None and jira_number is not None
            workspace_dirname = f"{repo_name}-{jira_number}"
            full_path = base_path / user_name / workspace_dirname
            logger.info(f"Constructed path with in if full_path {str(full_path)}")
        else:
            full_path = base_path
            logger.info(f"Constructed path with in if full_path {str(full_path)}")

        # Resolve to absolute path
        full_path = full_path.resolve()
        return full_path

    def read_file(
        self,
        *,
        file_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bytes:
        """
        Read file content from the local filesystem.

        Args:
            file_name: Relative path to the file (e.g., "folder/subfolder/file.xlsx")
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bytes: File content as bytes

        Raises:
            Exception: If file cannot be read
        """
        try:
            full_path = os.path.join(
                self._get_base_path(
                    user_name=user_name, repo_name=repo_name, jira_number=jira_number
                ),
                file_name,
            )
            with open(full_path, "rb") as f:
                return f.read()
        except Exception as e:
            raise Exception(f"Error reading file from local filesystem: {e}")

    def write_file(
        self,
        *,
        file_name: str,
        file_content: bytes,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Write file content to the local filesystem.

        Args:
            file_name: Relative path to the file (e.g., "folder/subfolder/file.xlsx")
            file_content: File content as bytes
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful

        Raises:
            Exception: If file cannot be written
        """
        try:
            full_path = os.path.join(
                self._get_base_path(
                    user_name=user_name, repo_name=repo_name, jira_number=jira_number
                ),
                file_name,
            )
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            with open(full_path, "wb") as f:
                f.write(file_content)
            return True
        except Exception as e:
            raise Exception(f"Error writing file to local filesystem: {e}")

    def list_files(
        self,
        *,
        path: str | None = None,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> List[str]:
        """
        List all files in the local filesystem, optionally filtered by path.

        Args:
            path: Optional path prefix to filter files (e.g., "folder/subfolder")
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            List[str]: List of relative file paths

        Raises:
            Exception: If files cannot be listed
        """
        try:
            if path is None:
                search_path = self._get_base_path(
                    user_name=user_name, repo_name=repo_name, jira_number=jira_number
                )
            else:
                search_path = os.path.join(
                    self._get_base_path(
                        user_name=user_name,
                        repo_name=repo_name,
                        jira_number=jira_number,
                    ),
                    path,
                )

            files = []
            for root, dirs, filenames in os.walk(search_path):
                for filename in filenames:
                    full_path = os.path.join(root, filename)
                    rel_path = os.path.relpath(
                        full_path,
                        self._get_base_path(
                            user_name=user_name,
                            repo_name=repo_name,
                            jira_number=jira_number,
                        ),
                    ).replace("\\", "/")
                    files.append(rel_path)
            return files
        except Exception as e:
            raise Exception(f"Error listing files from local filesystem: {e}")


class ExcelSheetsManager:
    """
    A comprehensive wrapper for Excel file operations with file server backend.

    Provides CRUD operations, worksheet management, and pandas DataFrame integration.
    Mirrors the GoogleSheetsManager API for easy swapping.
    """

    def __init__(self, base_url: str = "file_storage"):
        """
        Initialize the Excel Sheets manager.

        Args:
            base_url: Base URL of the file server API
        """
        # self.file_server = FileServerClient(base_url)
        self.file_server = LocalFileClient(base_dir=base_url)
        self._cache = {}  # In-memory cache for loaded workbooks
        print("✓ ExcelSheetsManager initialized")

    # Spreadsheet Discovery & Validation

    def validate_sheet_name(
        self,
        *,
        file_path: str,
        folder_path: str | None = None,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Check if Excel file exists and is accessible, optionally within a specific folder.

        Args:
            file_path: Name or path of the Excel file
            folder_path: Optional folder path to search within
            user_name: User name
            repo_name: Repository name
            jira_number: JIRA number

        Returns:
            bool: True if file exists and is accessible
        """

        try:
            full_path = self.find_spreadsheet_by_name(
                name=file_path,
                folder_path=folder_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return full_path is not None
        except Exception:
            return False

    def find_spreadsheet_by_name(
        self,
        *,
        name: str,
        folder_path: str | None = None,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> str:
        """
        Find an Excel file by name and return its full path, optionally within a specific folder.

        Args:
            name: Name of the Excel file (with or without .xlsx extension)
            folder_path: Optional folder path to search within
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            str: Full path to the Excel file

        Raises:
            ValueError: If file not found
        """
        try:
            # Ensure .xlsx extension
            if not name.endswith(".xlsx") and not name.endswith(".xls"):
                name = f"{name}.xlsx"

            # List all files
            all_files = self.file_server.list_files(
                path=folder_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            # Search for matching file
            for file_path in all_files:
                # Exact match
                if file_path == name:
                    print(f"✓ Found Excel file: {file_path}")
                    return file_path

                # Partial match (case-insensitive)
                if name.lower() in file_path.lower():
                    print(f"✓ Found Excel file: {file_path}")
                    return file_path

            raise ValueError(f"No Excel file found with name containing '{name}'")

        except Exception as e:
            raise Exception(f"Error finding Excel file: {e}")

    def _find_spreadsheet_in_folder(
        self,
        *,
        name: str,
        folder_path: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> str:
        """
        Find an Excel file by name within a specific folder.

        Args:
            name: Name of the Excel file
            folder_path: Folder path to search within
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            str: Full path to the Excel file
        """
        # Ensure .xlsx extension
        if not name.endswith(".xlsx") and not name.endswith(".xls"):
            name = f"{name}.xlsx"

        # Normalize folder path
        folder_path = folder_path.rstrip("/")

        # List files in folder
        all_files = self.file_server.list_files(
            path=folder_path,
            user_name=user_name,
            repo_name=repo_name,
            jira_number=jira_number,
        )

        for file_path in all_files:
            if file_path.startswith(folder_path):
                file_name = os.path.basename(file_path)
                if file_name == name or name.lower() in file_name.lower():
                    print(f"✓ Found Excel file: {file_path} in folder {folder_path}")
                    return file_path

        raise ValueError(f"No Excel file found with name '{name}' in folder {folder_path}")

    def get_spreadsheet_folder_info(
        self,
        *,
        file_path: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> dict:
        """
        Get folder information for an Excel file.

        Args:
            file_path: Full path to the Excel file
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            dict: Dictionary with file and folder information
        """
        try:
            # Extract folder path and file name
            folder_path = os.path.dirname(
                self.file_server._get_base_path(
                    user_name=user_name, repo_name=repo_name, jira_number=jira_number
                )
                / file_path
            )
            file_name = os.path.basename(
                self.file_server._get_base_path(
                    user_name=user_name, repo_name=repo_name, jira_number=jira_number
                )
                / file_path
            )

            return {
                "file_path": file_path,
                "file_name": file_name,
                "folder_path": folder_path,
            }

        except Exception as e:
            raise Exception(f"Error getting file info: {e}")

    def list_spreadsheets(
        self,
        *,
        folder_path: str | None = None,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> pd.DataFrame:
        """
        List all accessible Excel files as DataFrame.

        Args:
            folder_path: Optional folder path to filter files
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            pd.DataFrame: DataFrame with file information
        """
        try:
            all_files = self.file_server.list_files(
                path=folder_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            # Filter Excel files
            excel_files = [f for f in all_files if f.endswith((".xlsx", ".xls", ".xlsm"))]

            # Convert to DataFrame
            data = []
            for file_path in excel_files:
                data.append(
                    {
                        "file_path": file_path,
                        "file_name": os.path.basename(file_path),
                        "folder_path": os.path.dirname(file_path),
                    }
                )

            return pd.DataFrame(data)
        except Exception as e:
            raise Exception(f"Error listing Excel files: {e}")

    # Worksheet (Tab) Management

    def _load_workbook(
        self,
        *,
        file_path: str,
        use_cache: bool = True,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> openpyxl.Workbook:
        """
        Load Excel workbook from file server.

        Args:
            file_path: Full path to the Excel file
            use_cache: Whether to use cached workbook
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            openpyxl.Workbook: Loaded workbook object
        """
        if use_cache and file_path in self._cache:
            return self._cache[file_path]

        # Read file from server
        file_content = self.file_server.read_file(
            file_name=file_path,
            user_name=user_name,
            repo_name=repo_name,
            jira_number=jira_number,
        )

        # Load workbook from bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_content))

        if use_cache:
            self._cache[file_path] = wb

        return wb

    def _save_workbook(
        self,
        *,
        file_path: str,
        wb: "openpyxl.Workbook",
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Save Excel workbook to file server.

        Args:
            file_path: Full path to the Excel file
            wb: Workbook object to save
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        # Save workbook to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.read()

        # Write to file server
        result = self.file_server.write_file(
            file_name=file_path,
            file_content=file_content,
            user_name=user_name,
            repo_name=repo_name,
            jira_number=jira_number,
        )

        # Update cache
        self._cache[file_path] = wb

        return result

    def list_worksheets(
        self,
        *,
        file_path: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> List[str]:
        """
        Get all worksheet names.

        Args:
            file_path: Full path to the Excel file
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            List[str]: List of worksheet names
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return wb.sheetnames
        except Exception as e:
            raise Exception(f"Error listing worksheets: {e}")

    def validate_worksheet_name(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Check if worksheet exists.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if worksheet exists
        """
        try:
            worksheets = self.list_worksheets(file_path=file_path)
            return worksheet_name in worksheets
        except Exception:
            return False

    def create_worksheet(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        rows: int = 1000,
        cols: int = 26,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Create new worksheet.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name for the new worksheet
            rows: Number of rows (not used in openpyxl, kept for API compatibility)
            cols: Number of columns (not used in openpyxl, kept for API compatibility)
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            wb.create_sheet(title=worksheet_name)
            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True
        except Exception as e:
            print(f"Error creating worksheet: {e}")
            return False

    def delete_worksheet(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Delete worksheet.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet to delete
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            if worksheet_name in wb.sheetnames:
                del wb[worksheet_name]
                self._save_workbook(
                    file_path=file_path,
                    wb=wb,
                    user_name=user_name,
                    repo_name=repo_name,
                    jira_number=jira_number,
                )
                return True
            return False
        except Exception as e:
            print(f"Error deleting worksheet: {e}")
            return False

    def rename_worksheet(
        self,
        *,
        file_path: str,
        old_name: str,
        new_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Rename worksheet.

        Args:
            file_path: Full path to the Excel file
            old_name: Current worksheet name
            new_name: New worksheet name
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            if old_name in wb.sheetnames:
                wb[old_name].title = new_name
                self._save_workbook(
                    file_path=file_path,
                    wb=wb,
                    user_name=user_name,
                    repo_name=repo_name,
                    jira_number=jira_number,
                )
                return True
            return False
        except Exception as e:
            print(f"Error renaming worksheet: {e}")
            return False

    def duplicate_worksheet(
        self,
        *,
        file_path: str,
        source_worksheet: str,
        new_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Duplicate worksheet.

        Args:
            file_path: Full path to the Excel file
            source_worksheet: Name of worksheet to duplicate
            new_name: Name for the duplicated worksheet
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            if source_worksheet in wb.sheetnames:
                source_ws = wb[source_worksheet]
                target_ws = wb.copy_worksheet(source_ws)
                target_ws.title = new_name
                self._save_workbook(
                    file_path=file_path,
                    wb=wb,
                    user_name=user_name,
                    repo_name=repo_name,
                    jira_number=jira_number,
                )
                return True
            return False
        except Exception as e:
            print(f"Error duplicating worksheet: {e}")
            return False

    # Spreadsheet Copy Operations

    def copy_spreadsheet_to_folder(
        self,
        *,
        source_file_path: str,
        destination_folder_path: str,
        new_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> str:
        """
        Copy an Excel file to a specific folder with a new name.

        Args:
            source_file_path: The path of the file to copy
            destination_folder_path: The path of the destination folder
            new_name: The new name for the copied file
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            str: The path of the newly created file

        Raises:
            Exception: If the copy operation fails
        """
        try:
            # Ensure .xlsx extension
            if not new_name.endswith(".xlsx") and not new_name.endswith(".xls"):
                new_name = f"{new_name}.xlsx"

            # Normalize folder path
            destination_folder_path = destination_folder_path.rstrip("/")

            # Construct destination path
            destination_path = f"{destination_folder_path}/{new_name}"

            print(f"📋 Copying Excel file from: {source_file_path}...")

            # Read source file
            file_content = self.file_server.read_file(
                file_name=source_file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            # Write to destination
            self.file_server.write_file(
                file_name=destination_path,
                file_content=file_content,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            print("✓ Successfully copied Excel file to folder")
            print(f"  New file path: {destination_path}")

            return destination_path

        except Exception as e:
            raise Exception(f"Error copying Excel file to folder: {e}")

    def copy_spreadsheet_from_path(
        self,
        *,
        source_path: str,
        destination_folder_path: str,
        new_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> str:
        """
        Copy an Excel file from a path to a specific folder.

        Args:
            source_path: The full path to the source file
            destination_folder_path: The path of the destination folder
            new_name: The new name for the copied file
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            str: The path of the newly created file
        """
        return self.copy_spreadsheet_to_folder(
            source_file_path=source_path,
            destination_folder_path=destination_folder_path,
            new_name=new_name,
            user_name=user_name,
            repo_name=repo_name,
            jira_number=jira_number,
        )

    # Read Operations

    def get_sheet_data(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        range: str | None = None,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> pd.DataFrame:
        """
        Read data as DataFrame.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            range: Optional Excel range (e.g., "A1:D10")
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            pd.DataFrame: Data from the worksheet
        """
        try:
            # Read file from server
            file_content = self.file_server.read_file(
                file_name=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            if range:
                # Parse range and use usecols/skiprows
                # This is a simplified implementation
                df = pd.read_excel(io.BytesIO(file_content), sheet_name=worksheet_name)
                return self._extract_range_from_dataframe(df=df, range=range)
            else:
                # Get all data
                df = pd.read_excel(io.BytesIO(file_content), sheet_name=worksheet_name)
                return df

        except Exception as e:
            raise Exception(f"Error getting sheet data: {e}")

    def get_cell_value(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        row: int,
        col: int,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> str:
        """
        Read single cell.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            str: Cell value
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]
            cell = ws.cell(row=row, column=col)
            return str(cell.value) if cell.value is not None else ""
        except Exception as e:
            raise Exception(f"Error getting cell value: {e}")

    def get_range_values(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        range: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> pd.DataFrame:
        """
        Read specific range as DataFrame.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            range: Excel range (e.g., "A1:D10")
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            pd.DataFrame: Data from the range
        """
        try:
            return self.get_sheet_data(
                file_path=file_path,
                worksheet_name=worksheet_name,
                range=range,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
        except Exception as e:
            raise Exception(f"Error getting range values: {e}")

    # Write Operations

    def append_rows(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        df: pd.DataFrame,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Append DataFrame rows to end.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            df: DataFrame to append
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Check if worksheet is empty
            max_row = ws.max_row
            has_data = False

            if max_row > 0:
                # Check if there's any non-empty cell
                for row in ws.iter_rows(min_row=1, max_row=max_row):
                    if any(cell.value for cell in row):
                        has_data = True
                        break

            # Determine starting row
            if not has_data:
                # Empty worksheet - include headers
                start_row = 1
                include_header = True
            else:
                # Worksheet has data - append after last row, no headers
                start_row = max_row + 1
                include_header = False

            # Append data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=include_header)):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=start_row + r_idx, column=c_idx, value=value)

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error appending rows: {e}")
            return False

    def update_cell(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        row: int,
        col: int,
        value: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Update single cell.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            value: Value to set
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]
            ws.cell(row=row, column=col, value=value)
            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True
        except Exception as e:
            print(f"Error updating cell: {e}")
            return False

    def update_range(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        range: str,
        df: pd.DataFrame,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Update range with DataFrame.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            range: Excel range (e.g., "A1:D10")
            df: DataFrame to write
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Parse range to get starting cell
            start_cell = range.split(":")[0]
            start_row, start_col = self._parse_cell_reference(start_cell)

            # Write DataFrame to range
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row, start=0):
                    ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error updating range: {e}")
            return False

    def insert_rows(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        start_row: int,
        df: pd.DataFrame,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Insert DataFrame rows at position.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            start_row: Row number to insert at (1-indexed)
            df: DataFrame to insert
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Insert empty rows
            ws.insert_rows(start_row, amount=len(df))

            # Write DataFrame data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=start_row + r_idx, column=c_idx, value=value)

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error inserting rows: {e}")
            return False

    def clear_range(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        range: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Clear data in range.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            range: Excel range (e.g., "A1:D10")
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Clear range
            for row in ws[range]:
                for cell in row:
                    cell.value = None

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error clearing range: {e}")
            return False

    # Delete Operations

    def delete_rows(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        start_row: int,
        end_row: int,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Delete row range.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            start_row: Starting row number (1-indexed)
            end_row: Ending row number (1-indexed, inclusive)
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Calculate amount of rows to delete
            amount = end_row - start_row + 1
            ws.delete_rows(start_row, amount)

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error deleting rows: {e}")
            return False

    def delete_columns(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        start_col: int,
        end_col: int,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Delete column range.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            start_col: Starting column number (1-indexed)
            end_col: Ending column number (1-indexed, inclusive)
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Calculate amount of columns to delete
            amount = end_col - start_col + 1
            ws.delete_cols(start_col, amount)

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error deleting columns: {e}")
            return False

    # Utility Functions

    def format_as_markdown(self, df: pd.DataFrame, sheet_name: str | None = None) -> str:
        """
        Convert DataFrame to markdown (keep for compatibility).

        Args:
            df: DataFrame to convert
            sheet_name: Optional sheet name for header

        Returns:
            str: Markdown formatted table
        """
        if df.empty:
            return f"# {sheet_name or 'Sheet'}\n\n*No data found*"

        # Create header
        markdown = f"# {sheet_name or 'Sheet'}\n\n"

        # Generate markdown table
        markdown += df.to_markdown(index=False)

        return markdown

    def _get_column_letter(self, col_index: int) -> str:
        """
        Convert column index to letter (A, B, C...).

        Args:
            col_index: Column index (1-indexed)

        Returns:
            str: Column letter
        """
        return get_column_letter(col_index)

    def _parse_range(self, range: str) -> Tuple[str, int, int, int, int]:
        """
        Parse A1 notation range.

        Args:
            range: Excel range (e.g., "A1:D10")

        Returns:
            Tuple with range components
        """
        # Simple implementation - can be enhanced
        if ":" in range:
            start_cell, end_cell = range.split(":")
            start_row, start_col = self._parse_cell_reference(start_cell)
            end_row, end_col = self._parse_cell_reference(end_cell)
            return range, start_row, start_col, end_row, end_col
        else:
            row, col = self._parse_cell_reference(range)
            return range, row, col, row, col

    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """
        Parse cell reference like "A1" into (row, col).

        Args:
            cell_ref: Cell reference (e.g., "A1", "B10")

        Returns:
            Tuple[int, int]: (row, column) both 1-indexed
        """
        from openpyxl.utils import coordinate_to_tuple

        row, col = coordinate_to_tuple(cell_ref)

        return row, col

    def _extract_range_from_dataframe(self, *, df: pd.DataFrame, range: str) -> pd.DataFrame:
        """
        Extract a specific range from a DataFrame.

        Args:
            df: Source DataFrame
            range: Excel range (e.g., "A1:D10")

        Returns:
            pd.DataFrame: Extracted data
        """
        # Parse range
        _, start_row, start_col, end_row, end_col = self._parse_range(range)

        # Adjust for 0-indexed DataFrame (assuming row 1 is header)
        start_row_idx = max(0, start_row - 1)
        end_row_idx = end_row
        start_col_idx = start_col - 1
        end_col_idx = end_col

        # Extract range
        result = df.iloc[start_row_idx:end_row_idx, start_col_idx:end_col_idx]

        return result

    def _dataframe_to_values(self, df: pd.DataFrame) -> List[List]:
        """
        Convert DataFrame to list of lists format.

        Args:
            df: DataFrame to convert

        Returns:
            List[List]: Data with headers
        """
        # Include headers
        values = [df.columns.tolist()]
        values.extend(df.values.tolist())
        return values

    def _values_to_dataframe(self, values: List[List], headers: bool = True) -> pd.DataFrame:
        """
        Convert list of lists format to DataFrame.

        Args:
            values: Data as list of lists
            headers: Whether first row contains headers

        Returns:
            pd.DataFrame: Converted DataFrame
        """
        if not values:
            return pd.DataFrame()

        if headers and len(values) > 1:
            df = pd.DataFrame(values[1:], columns=values[0])
        else:
            df = pd.DataFrame(values)

        return df

    def clear_worksheet(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Clear all data from a worksheet.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            wb = self._load_workbook(
                file_path=file_path,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            ws = wb[worksheet_name]

            # Clear all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            self._save_workbook(
                file_path=file_path,
                wb=wb,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )
            return True

        except Exception as e:
            print(f"Error clearing worksheet: {e}")
            return False

    def upsert_sheet_data(
        self,
        *,
        file_path: str,
        worksheet_name: str,
        df: pd.DataFrame,
        user_name: str | None = None,
        repo_name: str | None = None,
        jira_number: str | None = None,
    ) -> bool:
        """
        Upsert data into a worksheet. If the worksheet exists, it will be cleared first.
        If the worksheet doesn't exist, it will be created with the provided data.

        Args:
            file_path: Full path to the Excel file
            worksheet_name: Name of the worksheet
            df: DataFrame to upsert
            user_name: User name (optional)
            repo_name: Repository name (optional)
            jira_number: JIRA number (optional)

        Returns:
            bool: True if successful
        """
        try:
            # Check if worksheet exists
            worksheet_exists = self.validate_worksheet_name(
                file_path=file_path,
                worksheet_name=worksheet_name,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            if worksheet_exists:
                # Clear existing worksheet
                print(f"Worksheet '{worksheet_name}' exists, clearing it first")
                clear_success = self.clear_worksheet(
                    file_path=file_path,
                    worksheet_name=worksheet_name,
                    user_name=user_name,
                    repo_name=repo_name,
                    jira_number=jira_number,
                )
                if not clear_success:
                    raise Exception(f"Failed to clear worksheet '{worksheet_name}'")
            else:
                # Create new worksheet
                print(f"Worksheet '{worksheet_name}' doesn't exist, creating it")
                create_success = self.create_worksheet(
                    file_path=file_path,
                    worksheet_name=worksheet_name,
                    rows=1000,
                    cols=26,
                    user_name=user_name,
                    repo_name=repo_name,
                    jira_number=jira_number,
                )
                if not create_success:
                    raise Exception(f"Failed to create worksheet '{worksheet_name}'")

            # Insert data into the worksheet
            success = self.append_rows(
                file_path=file_path,
                worksheet_name=worksheet_name,
                df=df,
                user_name=user_name,
                repo_name=repo_name,
                jira_number=jira_number,
            )

            if success:
                action = "cleared and updated" if worksheet_exists else "created with data"
                print(f"✓ Worksheet '{worksheet_name}' {action} successfully")
            else:
                print(f"Failed to insert data into worksheet '{worksheet_name}'")

            return success

        except Exception as e:
            print(f"Error upserting sheet data: {e}")
            return False

    def clear_cache(self):
        """Clear the in-memory workbook cache."""
        self._cache.clear()
        print("✓ Workbook cache cleared")


if __name__ == "__main__":
    # Example usage
    # manager = ExcelSheetsManager(base_dir="path/to/data")
    manager = ExcelSheetsManager(base_url="file_storage")
    # Add your test code here if needed
    resp = manager.get_sheet_data(file_path="Models.xlsx", worksheet_name="PaymentOrder")
    # print resp as formatted json
    print(resp.to_json(orient="records", indent=2))

    resp = manager.get_sheet_data(
        file_path="Models.xlsx",
        worksheet_name="PaymentOrder",
        user_name="pralhad",
        repo_name="myrepo",
        jira_number="JIRA-1",
    )
    # print resp as formatted json
    print(resp.to_json(orient="records", indent=2))
